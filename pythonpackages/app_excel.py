import openpyxl

#wb= openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

wb.create_sheet("Sheet2",0)

cell = sheet["a2"]
print("\n value",cell.value)
print( "\n row",cell.row)
print( "\n column",cell.column)
print("\n coordinate ",cell.coordinate)

cell = sheet.cell(row=1,column=1)
print(sheet.max_row)
print(sheet.max_column)

print("\n iterate through values in a spreadsheet \n")
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)


column = sheet["a"]
cells= sheet["a:c"]
print(" print column \n", column)

sheet[1:3]
sheet = wb["Sheet1"]
sheet.append([1,2,3])

print("\n iterate through values in a spreadsheet \n")
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)
wb.save("transactions2.xlsx")

for row in range(1,10):
    cell = sheet.cell(row,1)
    print(cell.value)

sheet.append([1,2,3])

#sheet.append([1,2,3])


#cell.value = 1
#wb.remove_sheet(sheet)
